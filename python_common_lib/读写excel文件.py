"""
@Time:2022/1/8 13:48
@Author:deanywang
@File:读写excel文件.py
@return:""
"""
import os
from openpyxl import load_workbook

"""
openpyxl:  .xlsx读写操作，可支持以往格式excel，
优点： xlrd、xlwt过时
测试数据：事先写好在excel
平常操作excel的流程（3个对象）
    工作簿（Workbook）
    表单（Sheet）
    单元格（Cell）


1、准备数据
2、load_workbook模块，去打开测试数据文件，生成WorkBook对象（wb）
3、根据表单名称选择表单（sh）：wb['表单名称']
4、在表单中，获取单元格的数据
    4.1 单元格对象：sh.cell(row,colum)  #下标从1开始
    4.2 .value获取单元格的值
    4.3 修改数据：sh.cell（row,colum）.value=新的值

5、获取当前表单中总行数和总列数
    sh.max_row  #总行数
    sh.max_colum  #总列数

6、在表单中，修改单元格的数据
7、保存数据（保存整个工作簿）
    WorkBook对象（wb.save(文件路径)）
    保存到原文件的时候需要注意，需要文件没有被占用，否则会报权限不允许错误

8、按行读取数据：
    sh.rows=所有行的数据
    list(sh.rows) ，每一行是个元祖，元祖里面放的是每一行的单元格
"""


file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_case.xlsx")
print(file_path)

# 2、load_workbook模块，去打开测试数据文件，生成WorkBook对象（wb）
wb = load_workbook(file_path)

# 3、根据表单名称选择表单（sh）：wb['表单名称']
sh = wb["login"]
print(sh)

# 4、在表单中，获取单元格的数据
cel = sh.cell(2, 2).value
print(cel)

# 5、获取当前表单中总行数和总列数
rows = sh.max_row
colums = sh.max_column
print(rows, colums)

# 6、在表单中，修改单元格的数据
sh.cell(2, 2).value = '1111'

# 7、保存数据（保存整个工作簿）
wb.save(file_path)
wb.close()

# 8、将表格的数据组装成列表
# 获取用例标题
titles = []
for title in list(sh.rows)[0]:
    print(title.value)
    titles.append(title.value)
print(titles)

# 获取用例数据方式1：
# data=[]
# for item in  list(sh.rows)[1:]:   #过滤第一行
#     data_dict={}
#     for index in range(len(item)):
#         print(index,item[index],item[index].value)
#         data_dict[titles[index]]=item[index].value
#     data.append(data_dict)
# print(data)

# 获取用例数据方式2：
datas = []
for item in list(sh.rows)[1:]:  # 过滤第一行
    dd = []
    for cel in item:
        value = cel.value
        dd.append(value)
    print(dd)
    data = dict(zip(titles, dd))
    data['pwd'] = eval(data['pwd'])
    datas.append(data)
print(datas)
